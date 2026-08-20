from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import load_workbook

from .api import DJENClient
from .classify import classify
from .excel import create_report
from .normalize import normalize_item
from .storage import PublicationStore
from .time_utils import brasilia_now


def run_packaged_self_test() -> None:
    # Garante que certifi/SSL tambem foram empacotados, sem fazer chamada de rede.
    client = DJENClient(min_interval=0)
    assert client._ssl_context is not None

    raw = {
        "id": 1,
        "hash": "selftest",
        "numero_processo": "00000000020268000000",
        "siglaTribunal": "TJXX",
        "data_disponibilizacao": "2026-08-19",
        "tipoComunicacao": "Intimacao",
        "texto": "Teste interno do pacote.",
        "destinatarioadvogados": [
            {"advogado": {"nome": "TESTE", "numero_oab": "123456", "uf_oab": "PR"}}
        ],
    }
    pub = classify(
        normalize_item(
            raw, "123456", "PR", collected_at=brasilia_now(), target_name="João de Sêne"
        )
    )
    with tempfile.TemporaryDirectory(prefix="djen-selftest-") as tmp:
        root = Path(tmp)
        with PublicationStore(root / "teste.sqlite3") as store:
            assert store.upsert(pub) == "NOVA"
            assert store.upsert(pub) == "JA_CONHECIDA"
        execution = {
            "started_at": brasilia_now().isoformat(timespec="seconds"),
            "finished_at": brasilia_now().isoformat(timespec="seconds"),
            "start_date": "2026-08-19",
            "end_date": "2026-08-19",
            "oabs": "João de Sêne (123456/PR)",
            "total_raw": 1,
            "total_normalized": 1,
            "total_new": 1,
            "total_updated": 0,
            "requests_made": 0,
            "complete": True,
            "duration_seconds": 0,
            "source": "selftest",
            "error": "",
        }
        path = create_report([pub], [pub], execution, root)
        wb = load_workbook(path, read_only=True)
        assert wb.sheetnames[0] == "RESUMO"
        assert "NOVAS_PUBLICACOES" in wb.sheetnames
        assert "CONFIGURACAO_E_EXECUCAO" not in wb.sheetnames
        assert wb["NOVAS_PUBLICACOES"]["I5"].value == "João de Sêne (123456/PR)"
        assert wb["NOVAS_PUBLICACOES"]["A4"].value == "Classificação"
        wb.close()


def run_stable_copy_self_test() -> None:
    import hashlib
    import os
    import sys

    if os.name != "nt" or not bool(getattr(sys, "frozen", False)):
        raise RuntimeError("Este autoteste exige o executável Windows empacotado.")

    from .scheduler import _stable_command

    command = _stable_command()
    target = Path(command[0])
    source = Path(sys.executable)
    if command[1:] != ["--automatico"]:
        raise RuntimeError("A cópia agendada não recebeu --automatico.")
    if not target.exists() or target.stat().st_size == 0:
        raise RuntimeError("A cópia estável do executável nao foi criada.")

    def sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    if sha256(source) != sha256(target):
        raise RuntimeError("A cópia estável difere do executável da Release.")
