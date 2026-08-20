from __future__ import annotations

import os
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from .normalize import Publication, strip_html
from .paths import fallback_reports_dir, reports_dir
from .time_utils import brasilia_now

EXCEL_CELL_LIMIT_SAFE = 30000
ILLEGAL_XML_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

# Paleta neutra/profissional. O arquivo é público e não depende de identidade visual
# de nenhum escritório específico.
NAVY = "17324D"
NAVY_2 = "234A68"
BLUE = "2F6B9A"
GOLD = "C59A3D"
WHITE = "FFFFFF"
TEXT = "24313C"
MUTED = "66737F"
LIGHT_BG = "F5F7F9"
LIGHT_BLUE = "EAF2F8"
BORDER = "D9E1E7"
GREEN = "E8F4EA"
GREEN_TEXT = "276738"
AMBER = "FFF3CD"
AMBER_TEXT = "7A5A00"
RED = "FDE8E7"
RED_TEXT = "9C2C2C"
GRAY = "EEF1F3"
GRAY_TEXT = "4F5B64"
UPDATE_BLUE = "E7F1FA"
UPDATE_TEXT = "245D8C"

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER),
    right=Side(style="thin", color=BORDER),
    top=Side(style="thin", color=BORDER),
    bottom=Side(style="thin", color=BORDER),
)

# Colunas que o usuário realmente precisa enxergar ao abrir a planilha.
VISIBLE_COLUMNS = [
    ("classificacao", "Classificação", 18),
    ("situacao_coleta", "Situação", 16),
    ("link_oficial", "Inteiro teor", 18),
    ("data_disponibilizacao", "Disponibilização", 16),
    ("numero_processo_formatado", "Processo", 25),
    ("sigla_tribunal", "Tribunal", 12),
    ("orgao_julgador", "Órgão julgador", 36),
    ("tipo_comunicacao", "Comunicação", 20),
    ("oab_monitorada", "OAB monitorada", 28),
    ("nome_advogado", "Advogado(s) na publicação", 34),
    ("partes", "Partes", 40),
    ("motivo_classificacao", "Motivo da classificação", 44),
    ("texto_visual", "Texto da publicação", 60),
    ("link_consulta_djen", "Consulta no DJEN", 19),
]

# Dados técnicos continuam integralmente no arquivo, mas ficam ocultos por padrão
# para a planilha não parecer uma exportação crua de banco de dados.
TECH_COLUMNS = [
    ("tribunal", "Tribunal - nome completo"),
    ("data_publicacao", "Data de publicação"),
    ("tipo_documento", "Tipo de documento"),
    ("classe_processual", "Classe processual"),
    ("meio", "Meio - código"),
    ("meio_completo", "Meio - descrição"),
    ("oab", "OAB retornada pela fonte"),
    ("uf_oab", "UF da OAB retornada"),
    ("oab_consultada", "OAB consultada"),
    ("uf_consultada", "UF consultada"),
    ("nome_oab_consultada", "Nome/apelido configurado"),
    ("rotulo_oab_consultada", "Rótulo local da OAB"),
    ("source_id", "Identificador da comunicação"),
    ("source_hash", "Hash da comunicação"),
    ("motivo_cancelamento", "Motivo de cancelamento"),
    ("status_fonte", "Status na fonte"),
    ("ativo_fonte", "Ativo na fonte"),
    ("coletado_em", "Data/hora da coleta"),
    ("texto_integral_1", "Texto integral - parte 1"),
    ("numero_processo_original", "Número do processo - valor original"),
    ("url_inteiro_teor", "URL do inteiro teor"),
    ("url_consulta_djen", "URL da consulta no DJEN"),
    ("classificacao_codigo", "Código da classificação"),
]

SHEET_META = {
    "NOVAS_PUBLICACOES": ("NOVAS PUBLICAÇÕES", "Publicações novas ou alteradas desde a última execução", NAVY),
    "TODAS_ENCONTRADAS": ("TODAS AS PUBLICAÇÕES ENCONTRADAS", "Tudo o que a consulta retornou na janela pesquisada", NAVY_2),
    "POSSIVEL_PRAZO": ("POSSÍVEL PRAZO", "Triagem automática para atenção. Não confirma juridicamente a existência de prazo.", "9E3B3B"),
    "REVISAR": ("REVISAR", "Itens incompletos, ambíguos ou que exigem conferência humana", "B27A16"),
    "ROTINA": ("ROTINA", "Itens provavelmente informativos, sem exclusão de revisão profissional", "477A54"),
}


def create_report(
    found: list[Publication],
    news: list[Publication],
    execution: dict,
    output_dir: Path | None = None,
) -> Path:
    explicit_output_dir = output_dir is not None
    output_dir = output_dir or reports_dir()
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = brasilia_now().strftime("%Y-%m-%d_%H%M%S")
    target = _unique_report_path(output_dir, timestamp)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    _write_summary_sheet(wb, found, news, execution)
    _write_publication_sheet(wb, "NOVAS_PUBLICACOES", news, execution)
    _write_publication_sheet(wb, "TODAS_ENCONTRADAS", found, execution)
    _write_publication_sheet(
        wb, "POSSIVEL_PRAZO", [p for p in news if p.classificacao == "POSSIVEL_PRAZO"], execution
    )
    _write_publication_sheet(wb, "REVISAR", [p for p in news if p.classificacao == "REVISAR"], execution)
    _write_publication_sheet(wb, "ROTINA", [p for p in news if p.classificacao == "ROTINA"], execution)

    try:
        wb.save(target)
    except PermissionError:
        target.unlink(missing_ok=True)
        retry_target = _unique_report_path(output_dir, f"{timestamp}_{os.getpid()}")
        try:
            wb.save(retry_target)
            target = retry_target
        except PermissionError:
            retry_target.unlink(missing_ok=True)
            if explicit_output_dir:
                raise
            fallback = fallback_reports_dir()
            target = _unique_report_path(fallback, timestamp)
            wb.save(target)
    except OSError:
        target.unlink(missing_ok=True)
        if explicit_output_dir:
            raise
        fallback = fallback_reports_dir()
        target = _unique_report_path(fallback, timestamp)
        wb.save(target)
    finally:
        wb.close()
    return target


def _unique_report_path(output_dir: Path, timestamp: str) -> Path:
    target = output_dir / f"DJEN_{timestamp}.xlsx"
    if not target.exists():
        return target
    counter = 2
    while True:
        candidate = output_dir / f"DJEN_{timestamp}_{counter}.xlsx"
        if not candidate.exists():
            return candidate
        counter += 1


def _write_summary_sheet(wb: Workbook, found: list[Publication], news: list[Publication], execution: dict) -> None:
    ws = wb.create_sheet(title="RESUMO")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 95
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = NAVY

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.merge_cells("A1:J2")
    title = ws["A1"]
    title.value = "DJEN MONITOR"
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.font = Font(name="Aptos Display", size=24, bold=True, color=WHITE)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:J3")
    subtitle = ws["A3"]
    subtitle.value = "Relatório de publicações do Diário de Justiça Eletrônico Nacional"
    subtitle.font = Font(name="Aptos", size=11, color=MUTED)
    subtitle.alignment = Alignment(vertical="center")
    ws.row_dimensions[3].height = 22

    possible = sum(1 for p in news if p.classificacao == "POSSIVEL_PRAZO")
    revisar = sum(1 for p in news if p.classificacao == "REVISAR")
    rotina = sum(1 for p in news if p.classificacao == "ROTINA")
    cards = [
        ("NOVAS", execution.get("total_new", 0), GREEN, GREEN_TEXT),
        ("ATUALIZADAS", execution.get("total_updated", 0), UPDATE_BLUE, UPDATE_TEXT),
        ("POSSÍVEL PRAZO", possible, RED, RED_TEXT),
        ("REVISAR", revisar, AMBER, AMBER_TEXT),
        ("TOTAL ENCONTRADO", execution.get("total_normalized", len(found)), LIGHT_BLUE, NAVY),
    ]
    for idx, (label, value, fill, font_color) in enumerate(cards):
        start_col = 1 + idx * 2
        end_col = start_col + 1
        cell_range = f"{get_column_letter(start_col)}5:{get_column_letter(end_col)}7"
        ws.merge_cells(cell_range)
        cell = ws.cell(5, start_col)
        cell.value = f"{label}\n{value}"
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Aptos Display", size=14, bold=True, color=font_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[5].height = 24
    ws.row_dimensions[6].height = 26
    ws.row_dimensions[7].height = 24

    complete = bool(execution.get("complete", False))
    status_fill = GREEN if complete else RED
    status_color = GREEN_TEXT if complete else RED_TEXT
    ws.merge_cells("A9:J9")
    status = ws["A9"]
    status.value = "COLETA COMPLETA" if complete else "ATENÇÃO: COLETA INCOMPLETA"
    status.fill = PatternFill("solid", fgColor=status_fill)
    status.font = Font(name="Aptos", size=11, bold=True, color=status_color)
    status.alignment = Alignment(horizontal="center", vertical="center")
    status.border = THIN_BORDER
    ws.row_dimensions[9].height = 26

    info_rows = [
        ("Período consultado", _period_text(execution)),
        ("OABs monitoradas", str(execution.get("oabs", ""))),
        ("Gerado em", _friendly_datetime(execution.get("finished_at")) or _friendly_datetime(execution.get("started_at"))),
        ("Requisições ao DJEN", execution.get("requests_made", "")),
        ("Fonte", execution.get("source", "")),
    ]
    row = 11
    for label, value in info_rows:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
        c1 = ws.cell(row, 1)
        c2 = ws.cell(row, 3)
        c1.value = label
        c2.value = safe_excel_text(value)
        c1.font = Font(name="Aptos", bold=True, color=NAVY)
        c2.font = Font(name="Aptos", color=TEXT)
        c1.fill = PatternFill("solid", fgColor=GRAY)
        c2.fill = PatternFill("solid", fgColor=WHITE)
        c1.border = c2.border = THIN_BORDER
        c1.alignment = Alignment(vertical="top")
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if label == "Fonte" and isinstance(value, str) and value.startswith("http"):
            c2.hyperlink = value
            c2.style = "Hyperlink"
        row += 1

    ws.merge_cells("A18:J20")
    notice = ws["A18"]
    notice.value = (
        "IMPORTANTE: as classificações são automáticas, provisórias e servem somente para triagem. "
        "Nenhuma publicação é descartada por baixa confiança. Revise as comunicações e consulte a fonte oficial."
    )
    notice.fill = PatternFill("solid", fgColor=AMBER)
    notice.font = Font(name="Aptos", size=10, bold=True, color=AMBER_TEXT)
    notice.alignment = Alignment(vertical="center", wrap_text=True)
    notice.border = THIN_BORDER

    links = [
        ("A22:B23", "Abrir novas publicações", "NOVAS_PUBLICACOES"),
        ("C22:D23", "Ver possíveis prazos", "POSSIVEL_PRAZO"),
        ("E22:F23", "Itens para revisar", "REVISAR"),
        ("G22:J23", "Itens de rotina", "ROTINA"),
    ]
    for cell_range, label, sheet_name in links:
        ws.merge_cells(cell_range)
        cell = ws[cell_range.split(":")[0]]
        cell.value = label
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.fill = PatternFill("solid", fgColor=NAVY_2)
        cell.font = Font(name="Aptos", bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _write_publication_sheet(wb: Workbook, title: str, rows: list[Publication], execution: dict) -> None:
    ws = wb.create_sheet(title=title)
    display_title, description, tab_color = SHEET_META[title]
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85

    max_chunks = max([1] + [len(split_text(p.texto_integral)) for p in rows])
    additional_text_columns = [
        (f"texto_integral_{i}", f"Texto integral - continuação {i}")
        for i in range(2, max_chunks + 1)
    ]
    columns = VISIBLE_COLUMNS + [(key, header, 16) for key, header in TECH_COLUMNS] + [
        (key, header, 18) for key, header in additional_text_columns
    ]
    visible_count = len(VISIBLE_COLUMNS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=visible_count)
    title_cell = ws.cell(1, 1)
    title_cell.value = display_title
    title_cell.fill = PatternFill("solid", fgColor=tab_color)
    title_cell.font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=visible_count)
    subtitle = ws.cell(2, 1)
    subtitle.value = (
        f"{description}  |  {len(rows)} registro(s)  |  Período: {_period_text(execution)}"
    )
    subtitle.fill = PatternFill("solid", fgColor=LIGHT_BG)
    subtitle.font = Font(name="Aptos", size=10, color=MUTED)
    subtitle.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    header_row = 4
    data_start = 5
    for col_idx, (_key, header, width) in enumerate(columns, start=1):
        cell = ws.cell(header_row, col_idx)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        if col_idx > visible_count:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True
    ws.row_dimensions[header_row].height = 34

    for pub in rows:
        chunks = split_text(pub.texto_integral)
        row_idx = ws.max_row + 1
        for col_idx, (key, _header, _width) in enumerate(columns, start=1):
            value = _publication_value(pub, key, chunks)
            cell = ws.cell(row_idx, col_idx)
            if key in {"data_disponibilizacao", "data_publicacao"}:
                parsed = _parse_iso_date(value)
                cell.value = parsed if parsed else safe_excel_text(value)
                if parsed:
                    cell.number_format = "dd/mm/yyyy"
            elif key == "link_oficial":
                _set_link_cell(cell, pub.link_oficial, "Abrir documento")
            elif key == "link_consulta_djen":
                _set_link_cell(cell, pub.link_consulta_djen, "Consultar DJEN")
            else:
                cell.value = safe_excel_text(value)

            if key not in {"link_oficial", "link_consulta_djen"}:
                cell.font = Font(name="Aptos", size=10, color=TEXT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

        _style_publication_row(ws, row_idx)
        ws.row_dimensions[row_idx].height = 76

    if rows:
        end_row = ws.max_row
        end_col = len(columns)
        table_name = {
            "NOVAS_PUBLICACOES": "tblNovasPublicacoes",
            "TODAS_ENCONTRADAS": "tblTodasEncontradas",
            "POSSIVEL_PRAZO": "tblPossivelPrazo",
            "REVISAR": "tblRevisar",
            "ROTINA": "tblRotina",
        }[title]
        table = Table(displayName=table_name, ref=f"A{header_row}:{get_column_letter(end_col)}{end_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    else:
        ws.merge_cells(start_row=data_start, start_column=1, end_row=data_start + 2, end_column=visible_count)
        empty = ws.cell(data_start, 1)
        empty.value = "Nenhuma publicação nesta categoria nesta execução."
        empty.fill = PatternFill("solid", fgColor=LIGHT_BG)
        empty.font = Font(name="Aptos", size=11, italic=True, color=MUTED)
        empty.alignment = Alignment(horizontal="center", vertical="center")
        empty.border = THIN_BORDER

    # Mantém apenas o cabeçalho congelado. Não congela colunas, evitando o divisor vertical no Excel.
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{max(header_row, ws.max_row)}"
    ws.sheet_view.selection[0].activeCell = "A5"
    ws.sheet_view.selection[0].sqref = "A5"


def _publication_value(pub: Publication, key: str, chunks: list[str]):
    if key == "oab_monitorada":
        return _display_monitored_oab(pub)
    if key == "numero_processo_formatado":
        return _format_process_number(pub.numero_processo)
    if key == "numero_processo_original":
        return pub.numero_processo
    if key == "texto_visual":
        return strip_html(pub.texto_integral)
    if key == "url_inteiro_teor":
        return pub.link_oficial
    if key == "url_consulta_djen":
        return pub.link_consulta_djen
    if key == "classificacao_codigo":
        return pub.classificacao
    if key.startswith("texto_integral_"):
        idx = int(key.rsplit("_", 1)[1]) - 1
        return chunks[idx] if idx < len(chunks) else ""
    if key == "classificacao":
        return _friendly_classification(pub.classificacao)
    if key == "situacao_coleta":
        return _friendly_collection_status(pub.situacao_coleta)
    return getattr(pub, key, "")


def _style_publication_row(ws, row_idx: int) -> None:
    class_cell = ws.cell(row_idx, 1)
    status_cell = ws.cell(row_idx, 2)
    classification = str(class_cell.value or "").upper()
    status = str(status_cell.value or "").upper()

    if "POSSÍVEL" in classification:
        class_cell.fill = PatternFill("solid", fgColor=RED)
        class_cell.font = Font(name="Aptos", size=10, bold=True, color=RED_TEXT)
    elif "REVISAR" in classification:
        class_cell.fill = PatternFill("solid", fgColor=AMBER)
        class_cell.font = Font(name="Aptos", size=10, bold=True, color=AMBER_TEXT)
    elif "ROTINA" in classification:
        class_cell.fill = PatternFill("solid", fgColor=GREEN)
        class_cell.font = Font(name="Aptos", size=10, bold=True, color=GREEN_TEXT)

    if "NOVA" in status:
        status_cell.fill = PatternFill("solid", fgColor=GREEN)
        status_cell.font = Font(name="Aptos", size=10, bold=True, color=GREEN_TEXT)
    elif "ATUALIZADA" in status:
        status_cell.fill = PatternFill("solid", fgColor=UPDATE_BLUE)
        status_cell.font = Font(name="Aptos", size=10, bold=True, color=UPDATE_TEXT)
    elif "SEM HISTÓRICO" in status:
        status_cell.fill = PatternFill("solid", fgColor=RED)
        status_cell.font = Font(name="Aptos", size=10, bold=True, color=RED_TEXT)
    else:
        status_cell.fill = PatternFill("solid", fgColor=GRAY)
        status_cell.font = Font(name="Aptos", size=10, bold=True, color=GRAY_TEXT)

    # Campos compactos ficam centralizados; texto jurídico permanece à esquerda.
    for col_idx in (1, 2, 3, 4, 6, 8, 14):
        ws.cell(row_idx, col_idx).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)


def _set_link_cell(cell, url: str, label: str) -> None:
    if isinstance(url, str) and re.match(r"^https?://", url):
        cell.value = label
        cell.hyperlink = url
        cell.font = Font(name="Aptos", size=10, bold=True, color="0563C1", underline="single")
    else:
        cell.value = "Não informado"
        cell.font = Font(name="Aptos", size=10, italic=True, color=MUTED)


def _display_monitored_oab(pub: Publication) -> str:
    if str(pub.rotulo_oab_consultada or "").strip():
        return str(pub.rotulo_oab_consultada).strip()
    oabs = _pipe_values(pub.oab_consultada)
    ufs = _pipe_values(pub.uf_consultada)
    names = _pipe_values(pub.nome_oab_consultada)
    if not oabs:
        return ""

    if len(ufs) == 1 and len(oabs) > 1:
        ufs = ufs * len(oabs)
    result = []
    for idx, number in enumerate(oabs):
        uf = ufs[idx] if idx < len(ufs) else (ufs[0] if ufs else "")
        registration = f"{number}/{uf}" if uf else number
        name = names[idx] if idx < len(names) else ""
        result.append(f"{name} ({registration})" if name else registration)
    return " | ".join(result)


def _pipe_values(value: str) -> list[str]:
    return [part.strip() for part in str(value or "").split(" | ") if part.strip()]


def _format_process_number(value: str) -> str:
    text = str(value or "").strip()
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) == 20:
        return f"{digits[:7]}-{digits[7:9]}.{digits[9:13]}.{digits[13]}.{digits[14:16]}.{digits[16:]}"
    return text


def _friendly_classification(value: str) -> str:
    return {
        "POSSIVEL_PRAZO": "POSSÍVEL PRAZO",
        "REVISAR": "REVISAR",
        "ROTINA": "ROTINA",
    }.get(str(value or ""), str(value or "").replace("_", " "))


def _friendly_collection_status(value: str) -> str:
    return {
        "NOVA": "NOVA",
        "ATUALIZADA": "ATUALIZADA",
        "JA_CONHECIDA": "JÁ CONHECIDA",
        "SEM_HISTORICO": "SEM HISTÓRICO",
    }.get(str(value or ""), str(value or "").replace("_", " "))


def _parse_iso_date(value) -> date | None:
    text = str(value or "").strip()
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return None
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _friendly_datetime(value) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return parsed.strftime("%d/%m/%Y %H:%M:%S")
    except ValueError:
        return text


def _period_text(execution: dict) -> str:
    start = _friendly_date(execution.get("start_date"))
    end = _friendly_date(execution.get("end_date"))
    if start and end:
        return f"{start} a {end}"
    return start or end or "Não informado"


def _friendly_date(value) -> str:
    parsed = _parse_iso_date(value)
    return parsed.strftime("%d/%m/%Y") if parsed else str(value or "")


def _yes_no(value) -> str:
    return "Sim" if bool(value) else "Não"


def split_text(value: str) -> list[str]:
    value = str(value or "")
    if not value:
        return [""]
    chunks: list[str] = []
    current: list[str] = []
    units = 0
    for char in value:
        char_units = len(char.encode("utf-16-le")) // 2
        if current and units + char_units > EXCEL_CELL_LIMIT_SAFE:
            chunks.append("".join(current))
            current = []
            units = 0
        current.append(char)
        units += char_units
    if current:
        chunks.append("".join(current))
    return chunks or [""]


def safe_excel_text(value) -> str | int | float | bool:
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value if value is not None else "")
    text = ILLEGAL_XML_CHARS_RE.sub(lambda m: f"\\x{ord(m.group(0)):02X}", text)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text
