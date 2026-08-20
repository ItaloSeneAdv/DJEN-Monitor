from datetime import date

from openpyxl import load_workbook

from djen_monitor.api import QueryResult
from djen_monitor.config import AppConfig, OABConfig
from djen_monitor.excel import create_report as real_create_report
from djen_monitor.runner import run_monitor
from djen_monitor.storage import PublicationStore


RAW = {
    "id": 777,
    "hash": "hash-777",
    "numero_processo": "00000000020268000000",
    "siglaTribunal": "TJXX",
    "data_disponibilizacao": "2026-08-19",
    "tipoComunicacao": "Intimacao",
    "texto": "Intimado para se manifestar.",
    "destinatarioadvogados": [
        {"advogado": {"nome": "ADVOGADO TESTE", "numero_oab": "123456", "uf_oab": "PR"}}
    ],
}


def test_two_successful_runs_generate_excel_and_deduplicate(tmp_path, monkeypatch):
    db = tmp_path / "dados.sqlite3"
    reports = tmp_path / "Planilhas"

    class FakeClient:
        def __init__(self, *args, **kwargs):
            self.requests_made = 0

        def query_oab(self, **kwargs):
            self.requests_made += 1
            return QueryResult(items=[RAW], announced_count=1, complete=True, requests_made=1)

    monkeypatch.setattr("djen_monitor.runner.DJENClient", FakeClient)
    monkeypatch.setattr("djen_monitor.runner.PublicationStore", lambda: PublicationStore(db))
    monkeypatch.setattr(
        "djen_monitor.runner.create_report",
        lambda found, news, execution: real_create_report(found, news, execution, reports),
    )
    monkeypatch.setattr("djen_monitor.runner.brasilia_today", lambda: date(2026, 8, 19))

    cfg = AppConfig(
        oabs=[OABConfig("123456", "PR")],
        janela_dias=3,
        consultar_variantes_oab=False,
        intervalo_requisicoes_segundos=0,
    )

    first = run_monitor(cfg)
    second = run_monitor(cfg)

    assert not first.error and not second.error
    assert first.found == 1 and first.new == 1
    assert second.found == 1 and second.new == 0
    assert first.report_path != second.report_path

    wb1 = load_workbook(first.report_path, read_only=True)
    wb2 = load_workbook(second.report_path, read_only=True)
    assert wb1["NOVAS_PUBLICACOES"]["B5"].value == "NOVA"
    assert wb2["NOVAS_PUBLICACOES"]["A5"].value == "Nenhuma publicação nesta categoria nesta execução."
    headers = [c.value for c in wb1["TODAS_ENCONTRADAS"][4]]
    assert "OAB retornada pela fonte" in headers
    assert "OAB consultada" in headers
    wb1.close()
    wb2.close()


def test_failure_in_one_oab_does_not_prevent_other_oabs(tmp_path, monkeypatch):
    from djen_monitor.api import DJENAPIError

    db = tmp_path / "multi.sqlite3"
    reports = tmp_path / "Relatorios"
    raw_ok = dict(RAW)
    raw_ok["id"] = 888
    raw_ok["destinatarioadvogados"] = [
        {"advogado": {"nome": "ADVOGADO DOIS", "numero_oab": "654321", "uf_oab": "SP"}}
    ]

    class PartialClient:
        def __init__(self, *args, **kwargs):
            self.requests_made = 0

        def query_oab(self, numero_oab, **kwargs):
            self.requests_made += 1
            if numero_oab == "123456":
                raise DJENAPIError("falha simulada")
            return QueryResult(items=[raw_ok], announced_count=1, complete=True, requests_made=1)

    monkeypatch.setattr("djen_monitor.runner.DJENClient", PartialClient)
    monkeypatch.setattr("djen_monitor.runner.PublicationStore", lambda: PublicationStore(db))
    monkeypatch.setattr(
        "djen_monitor.runner.create_report",
        lambda found, news, execution: real_create_report(found, news, execution, reports),
    )
    monkeypatch.setattr("djen_monitor.runner.brasilia_today", lambda: date(2026, 8, 19))

    cfg = AppConfig(
        oabs=[OABConfig("123456", "PR"), OABConfig("654321", "SP")],
        janela_dias=1,
        consultar_variantes_oab=False,
        intervalo_requisicoes_segundos=0,
    )
    result = run_monitor(cfg)
    assert result.found == 1
    assert result.new == 1
    assert "123456/PR" in result.error
    assert result.complete is False
    wb = load_workbook(result.report_path, read_only=True)
    assert wb["NOVAS_PUBLICACOES"]["B5"].value == "NOVA"
    wb.close()


def test_sqlite_failure_still_generates_emergency_excel(tmp_path, monkeypatch):
    reports = tmp_path / "Emergencia"

    class FakeClient:
        def __init__(self, *args, **kwargs):
            self.requests_made = 0

        def query_oab(self, **kwargs):
            self.requests_made += 1
            return QueryResult(items=[RAW], announced_count=1, complete=True, requests_made=1)

    class BrokenStore:
        def __init__(self, *args, **kwargs):
            raise OSError("banco indisponivel")

    monkeypatch.setattr("djen_monitor.runner.DJENClient", FakeClient)
    monkeypatch.setattr("djen_monitor.runner.PublicationStore", BrokenStore)
    monkeypatch.setattr(
        "djen_monitor.runner.create_report",
        lambda found, news, execution: real_create_report(found, news, execution, reports),
    )
    monkeypatch.setattr("djen_monitor.runner.brasilia_today", lambda: date(2026, 8, 19))

    cfg = AppConfig(
        oabs=[OABConfig("123456", "PR")],
        janela_dias=1,
        consultar_variantes_oab=False,
        intervalo_requisicoes_segundos=0,
    )
    result = run_monitor(cfg)
    assert result.complete is False
    assert "armazenamento local" in result.error
    assert result.report_path
    wb = load_workbook(result.report_path, read_only=True)
    assert wb["NOVAS_PUBLICACOES"]["B5"].value == "SEM HISTÓRICO"
    assert wb["NOVAS_PUBLICACOES"]["A5"].value == "REVISAR"  # classificação provisória
    wb.close()


def test_malformed_item_does_not_hide_later_items(tmp_path, monkeypatch):
    from djen_monitor.normalize import normalize_item as real_normalize_item

    db = tmp_path / "malformed.sqlite3"
    reports = tmp_path / "RelatoriosMalformados"
    bad = dict(RAW)
    bad["id"] = 1
    good = dict(RAW)
    good["id"] = 2
    good["hash"] = "hash-2"

    class FakeClient:
        def __init__(self, *args, **kwargs):
            self.requests_made = 0

        def query_oab(self, **kwargs):
            self.requests_made += 1
            return QueryResult(items=[bad, good], announced_count=2, complete=True, requests_made=1)

    def selective_normalize(item, *args, **kwargs):
        if item.get("id") == 1:
            raise ValueError("item quebrado")
        return real_normalize_item(item, *args, **kwargs)

    monkeypatch.setattr("djen_monitor.runner.DJENClient", FakeClient)
    monkeypatch.setattr("djen_monitor.runner.normalize_item", selective_normalize)
    monkeypatch.setattr("djen_monitor.runner.PublicationStore", lambda: PublicationStore(db))
    monkeypatch.setattr(
        "djen_monitor.runner.create_report",
        lambda found, news, execution: real_create_report(found, news, execution, reports),
    )
    monkeypatch.setattr("djen_monitor.runner.brasilia_today", lambda: date(2026, 8, 19))

    cfg = AppConfig(
        oabs=[OABConfig("123456", "PR")],
        janela_dias=1,
        consultar_variantes_oab=False,
        intervalo_requisicoes_segundos=0,
    )
    result = run_monitor(cfg)
    assert result.found == 1
    assert result.new == 1
    assert result.complete is False
    assert "comunicação 1" in result.error
    assert result.report_path


def test_search_window_catches_up_from_last_complete_execution(tmp_path, monkeypatch):
    db = tmp_path / "catchup.sqlite3"
    reports = tmp_path / "Catchup"
    with PublicationStore(db) as store:
        store.record_execution({
            "started_at": "2026-08-10T08:00:00-03:00",
            "finished_at": "2026-08-10T08:01:00-03:00",
            "start_date": "2026-08-08",
            "end_date": "2026-08-10",
            "total_raw": 0,
            "total_normalized": 0,
            "total_new": 0,
            "total_updated": 0,
            "requests_made": 1,
            "complete": True,
            "report_path": "",
            "error": "",
        })

    seen = {}

    class FakeClient:
        def __init__(self, *args, **kwargs):
            self.requests_made = 0
        def query_oab(self, numero_oab, uf_oab, start, end, **kwargs):
            seen["start"] = start
            seen["end"] = end
            self.requests_made += 1
            return QueryResult(items=[], announced_count=0, complete=True, requests_made=1)

    monkeypatch.setattr("djen_monitor.runner.DJENClient", FakeClient)
    monkeypatch.setattr("djen_monitor.runner.PublicationStore", lambda: PublicationStore(db))
    monkeypatch.setattr(
        "djen_monitor.runner.create_report",
        lambda found, news, execution: real_create_report(found, news, execution, reports),
    )
    monkeypatch.setattr("djen_monitor.runner.remember_last_report", lambda _path: None)
    monkeypatch.setattr("djen_monitor.runner.brasilia_today", lambda: date(2026, 8, 20))

    cfg = AppConfig(
        oabs=[OABConfig("123456", "PR")], janela_dias=3,
        consultar_variantes_oab=False, intervalo_requisicoes_segundos=0,
    )
    result = run_monitor(cfg)
    assert seen["start"] == date(2026, 8, 10)
    assert seen["end"] == date(2026, 8, 20)
    assert result.start_date == "2026-08-10"
