import json
from pathlib import Path

from openpyxl import load_workbook

from djen_monitor.classify import classify
from djen_monitor.excel import create_report
from djen_monitor.normalize import normalize_item

FIXTURE = Path(__file__).parent / "fixtures" / "comunicacoes_anonimizadas.json"


def execution_fixture(**overrides):
    data = {
        "app_version": "1.2.1",
        "started_at": "2026-08-19T08:00:00-03:00",
        "finished_at": "2026-08-19T08:00:10-03:00",
        "start_date": "2026-08-17",
        "end_date": "2026-08-19",
        "configured_start_date": "2026-08-17",
        "catchup_expanded": False,
        "oabs": "João da Silva (123456/PR)",
        "window_days": 3,
        "variants_enabled": True,
        "request_interval_seconds": 0.7,
        "total_raw": 1,
        "total_normalized": 1,
        "total_new": 1,
        "total_updated": 0,
        "requests_made": 1,
        "complete": True,
        "duration_seconds": 10,
        "source": "https://comunicaapi.pje.jus.br/api/v1/comunicacao",
        "error": "",
    }
    data.update(overrides)
    return data


def test_excel_has_polished_summary_expected_sheets_and_long_text(tmp_path):
    payload = json.loads(FIXTURE.read_text(encoding="utf-8"))
    pub = classify(normalize_item(payload["items"][0], "123456", "PR", target_name="João da Silva"))
    pub.texto_integral = "A" * 65000
    path = create_report([pub], [pub], execution_fixture(), tmp_path)

    wb = load_workbook(path, read_only=False)
    assert wb.sheetnames[0] == "RESUMO"
    assert "NOVAS_PUBLICACOES" in wb.sheetnames
    assert "POSSIVEL_PRAZO" in wb.sheetnames
    assert "CONFIGURACAO_E_EXECUCAO" not in wb.sheetnames
    assert wb["RESUMO"]["A1"].value == "DJEN MONITOR"

    ws = wb["NOVAS_PUBLICACOES"]
    headers = [cell.value for cell in ws[4]]
    assert "Texto da publicação" in headers
    assert "Texto integral - continuação 2" in headers
    assert "Texto integral - continuação 3" in headers
    assert ws["I5"].value == "João da Silva (123456/PR)"
    assert ws.freeze_panes == "A5"
    assert ws.tables

    # Colunas técnicas continuam no arquivo, mas ficam escondidas por padrão.
    technical_col = headers.index("Identificador da comunicação") + 1
    assert ws.column_dimensions[ws.cell(4, technical_col).column_letter].hidden is True
    wb.close()


def test_split_text_preserves_unicode_beyond_excel_utf16_limit():
    from djen_monitor.excel import EXCEL_CELL_LIMIT_SAFE, split_text

    original = "😀" * 20000 + "fim"
    chunks = split_text(original)
    assert "".join(chunks) == original
    assert len(chunks) >= 2
    assert all(len(chunk.encode("utf-16-le")) // 2 <= EXCEL_CELL_LIMIT_SAFE for chunk in chunks)


def test_excel_escapes_illegal_xml_control_chars(tmp_path):
    payload = json.loads(FIXTURE.read_text(encoding="utf-8"))
    pub = classify(normalize_item(payload["items"][0], "123456", "PR"))
    pub.texto_integral = "antes\x00depois\x0bfinal"
    path = create_report([pub], [pub], execution_fixture(), tmp_path)
    wb = load_workbook(path, read_only=True)
    ws = wb["NOVAS_PUBLICACOES"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=4, max_row=4))]
    row = [cell.value for cell in next(ws.iter_rows(min_row=5, max_row=5))]
    raw_text = row[headers.index("Texto integral - parte 1")]
    assert "\\x00" in raw_text and "\\x0B" in raw_text
    wb.close()


def test_excel_uses_friendly_links_and_accented_headers(tmp_path):
    payload = json.loads(FIXTURE.read_text(encoding="utf-8"))
    pub = classify(normalize_item(payload["items"][0], "123456", "PR", target_name="Márcia José"))
    path = create_report([pub], [pub], execution_fixture(oabs="Márcia José (123456/PR)"), tmp_path)
    wb = load_workbook(path, read_only=False)
    ws = wb["NOVAS_PUBLICACOES"]
    assert ws["A4"].value == "Classificação"
    assert ws["G4"].value == "Órgão julgador"
    assert ws["M4"].value == "Texto da publicação"
    assert ws["C4"].value == "Inteiro teor"
    assert ws["C5"].value in {"Abrir documento", "Não informado"}
    if ws["C5"].value == "Abrir documento":
        assert ws["C5"].hyperlink is not None
    wb.close()


def test_excel_visible_text_is_human_readable_and_raw_html_is_preserved(tmp_path):
    payload = json.loads(FIXTURE.read_text(encoding="utf-8"))
    pub = classify(normalize_item(payload["items"][0], "123456", "PR", target_name="João"))
    raw_html = "  <p>Olá <strong>mundo</strong>.</p>\n<p>Segunda linha.</p>  "
    pub.texto_integral = raw_html
    pub.numero_processo = "00005254920268160163"

    path = create_report([pub], [pub], execution_fixture(), tmp_path)
    wb = load_workbook(path, read_only=False)
    ws = wb["NOVAS_PUBLICACOES"]
    headers = [cell.value for cell in ws[4]]

    assert ws["E5"].value == "0000525-49.2026.8.16.0163"
    assert ws["M5"].value == "Olá mundo . Segunda linha."

    raw_col = headers.index("Texto integral - parte 1") + 1
    assert ws.cell(5, raw_col).value == raw_html
    assert ws.column_dimensions[ws.cell(4, raw_col).column_letter].hidden is True
    wb.close()


def test_excel_has_inteiro_teor_as_third_column_and_no_execution_sheet(tmp_path):
    payload = json.loads(FIXTURE.read_text(encoding="utf-8"))
    pub = classify(normalize_item(payload["items"][0], "123456", "PR"))
    path = create_report([pub], [pub], execution_fixture(), tmp_path)
    wb = load_workbook(path, read_only=False)
    assert "CONFIGURACAO_E_EXECUCAO" not in wb.sheetnames
    for sheet_name in ["NOVAS_PUBLICACOES", "TODAS_ENCONTRADAS", "POSSIVEL_PRAZO", "REVISAR", "ROTINA"]:
        ws = wb[sheet_name]
        assert ws["C4"].value == "Inteiro teor"
        assert ws.freeze_panes == "A5"
    wb.close()
